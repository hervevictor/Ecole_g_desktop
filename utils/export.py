from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, PageBreak
)
from PySide6.QtWidgets import QFileDialog, QMessageBox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from database.connexion import get_session
from database.models import ConfigurationEtablissement, PaiementEcolage


def _get_config():
    session = get_session()
    return session.query(ConfigurationEtablissement).first()


def export_eleves_pdf(eleves, parent=None):
    path, _ = QFileDialog.getSaveFileName(parent, "Enregistrer PDF", "liste_eleves.pdf", "PDF (*.pdf)")
    if not path:
        return
    try:
        config = _get_config()
        doc = SimpleDocTemplate(path, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elements = []

        nom_ecole = config.nom_etablissement if config else "École"
        elements.append(Paragraph(f"<b>{nom_ecole}</b>", styles['Title']))
        elements.append(Paragraph("Liste des Élèves", styles['Heading2']))
        elements.append(Spacer(1, 0.5*cm))

        data = [["Matricule", "Nom", "Prénom", "Classe", "Sexe", "Date Naiss.", "Parent", "Téléphone"]]
        for e in eleves:
            classe_str = str(e.nos_classe) if e.nos_classe else "—"
            ddn = e.date_de_naissance.strftime('%d/%m/%Y') if e.date_de_naissance else "—"
            data.append([
                e.matricule or '', e.nom, e.prenom, classe_str,
                e.sexe or '', ddn, e.nom_du_parent or '', str(e.tel_parent or '')
            ])

        col_widths = [2.5*cm, 3*cm, 3*cm, 3*cm, 1.5*cm, 3*cm, 4*cm, 2.5*cm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7fafc')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"Total : {len(eleves)} élève(s)", styles['Normal']))
        doc.build(elements)
        QMessageBox.information(parent, "Succès", f"PDF exporté : {path}")
        _ouvrir_fichier(path)
    except Exception as e:
        QMessageBox.critical(parent, "Erreur", f"Impossible de générer le PDF :\n{e}")


def export_eleves_excel(eleves, parent=None):
    path, _ = QFileDialog.getSaveFileName(parent, "Enregistrer Excel", "liste_eleves.xlsx", "Excel (*.xlsx)")
    if not path:
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Élèves"
        headers = ["Matricule", "Nom", "Prénom", "Classe", "Sexe",
                   "Date de naissance", "Lieu de naissance", "Nom parent", "Tél. parent", "Adresse"]
        header_fill = PatternFill("solid", fgColor="1A365D")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, e in enumerate(eleves, 2):
            classe_str = str(e.nos_classe) if e.nos_classe else "—"
            ddn = e.date_de_naissance.strftime('%d/%m/%Y') if e.date_de_naissance else "—"
            for col, val in enumerate([
                e.matricule or '', e.nom, e.prenom, classe_str, e.sexe or '',
                ddn, e.lieu_de_naissance or '', e.nom_du_parent or '',
                str(e.tel_parent or ''), e.adresse or ''
            ], 1):
                ws.cell(row=row, column=col, value=val)

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        wb.save(path)
        QMessageBox.information(parent, "Succès", f"Excel exporté : {path}")
        _ouvrir_fichier(path)
    except Exception as e:
        QMessageBox.critical(parent, "Erreur", f"Impossible d'exporter :\n{e}")


def export_historique_paiement_pdf(eleve, session, parent=None):
    if not eleve:
        return
    path, _ = QFileDialog.getSaveFileName(
        parent, "Enregistrer PDF",
        f"historique_{eleve.matricule}.pdf", "PDF (*.pdf)"
    )
    if not path:
        return
    try:
        config = _get_config()
        doc = SimpleDocTemplate(path, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        elements = []

        nom_ecole = config.nom_etablissement if config else "École"
        elements.append(Paragraph(f"<b>{nom_ecole}</b>", styles['Title']))
        elements.append(Paragraph(f"Historique des paiements — {eleve.nom} {eleve.prenom}", styles['Heading2']))
        elements.append(Paragraph(f"Matricule : {eleve.matricule}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))

        paiements = session.query(PaiementEcolage).filter_by(
            eleve_id=eleve.id, statut='valide'
        ).order_by(PaiementEcolage.date_paiement).all()

        data = [["Date", "Montant (FCFA)", "Mode", "Référence", "N° Reçu"]]
        total = 0
        for p in paiements:
            date_str = p.date_paiement.strftime('%d/%m/%Y') if p.date_paiement else '—'
            num_recu = p.recu.numero_recu if p.recu else '—'
            data.append([date_str, f"{p.montant:,}", p.mode_paiement or '—', p.reference or '—', num_recu])
            total += p.montant

        table = Table(data, colWidths=[3.5*cm, 4*cm, 3.5*cm, 4*cm, 3.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#276749')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fff4')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(f"<b>Total payé : {total:,} FCFA</b>", styles['Heading3']))

        du = eleve.montant_total_du(session)
        reste = du - total
        elements.append(Paragraph(f"Montant dû : {du:,} FCFA | Reste : {reste:,} FCFA", styles['Normal']))

        doc.build(elements)
        QMessageBox.information(parent, "Succès", f"PDF généré : {path}")
        _ouvrir_fichier(path)
    except Exception as e:
        QMessageBox.critical(parent, "Erreur", f"Erreur PDF :\n{e}")


def _ouvrir_fichier(path):
    import subprocess, sys
    try:
        if sys.platform == 'darwin':
            subprocess.Popen(['open', path])
        elif sys.platform == 'win32':
            os.startfile(path)
        else:
            subprocess.Popen(['xdg-open', path])
    except Exception:
        pass
